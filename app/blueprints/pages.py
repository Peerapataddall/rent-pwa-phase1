# app/blueprints/pages.py
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from sqlalchemy import or_

from app import db
from app.models import (
    Billing,
    Customer,
    Installment,
    Product,
    Quote,
    QuoteLine,
    Receipt,
    RentalContract,
    TaxInvoice,
    # ✅ Purchase / Inventory
    Asset,
    AssetPhoto,
    Brand,
    Category,
    GRN,
    GRNLine,
    PurchaseOrder,
    PurchaseOrderLine,
    StockMove,
    Supplier,
    Unit,
)

bp = Blueprint("pages", __name__)

from app.utils.contract_status import (  # noqa: E402
    auto_update_contract_status,
    ensure_contract_running_status,
    guard_contract_not_locked,
)

# =========================================================
# ✅ Thai Status Normalization (Phase Step 1 readiness)
# - เป้าหมายสุดท้าย: เก็บสถานะเป็น "คำไทย" ใน DB
# - แต่เพื่อไม่ให้ของเดิมพัง (ข้อมูลเก่า/โค้ดเก่า) เราจะรองรับทั้ง "อังกฤษเดิม" และ "ไทยใหม่"
# =========================================================

# Asset statuses (ไทยล้วน เป้าหมาย)
ASSET_STATUSES_TH = [
    "พร้อมใช้งาน",
    "จองแล้ว",
    "กำลังเช่า",
    "รอตรวจสภาพ",
    "ซ่อม",
    "สูญหาย",
    "จำหน่ายออก",
]

# Legacy -> Thai (รองรับข้อมูล/ฟอร์มเก่า)
ASSET_STATUS_LEGACY_TO_TH = {
    "AVAILABLE": "พร้อมใช้งาน",
    "RESERVED": "จองแล้ว",       # เผื่ออนาคต
    "RENTED": "กำลังเช่า",
    "INSPECT": "รอตรวจสภาพ",    # เผื่ออนาคต
    "REPAIR": "ซ่อม",
    "LOST": "สูญหาย",
    "RETIRED": "จำหน่ายออก",
}

# Thai -> Legacy candidates (ใช้ตอน filter ให้ match ข้อมูลเก่า)
ASSET_STATUS_TH_TO_LEGACY = {}
for k, v in ASSET_STATUS_LEGACY_TO_TH.items():
    ASSET_STATUS_TH_TO_LEGACY.setdefault(v, set()).add(k)


# PO statuses (ไทยล้วน เป้าหมาย)
PO_STATUSES_TH = ["ร่าง", "อนุมัติ", "รับบางส่วน", "รับครบ", "ยกเลิก"]

PO_STATUS_LEGACY_TO_TH = {
    "DRAFT": "ร่าง",
    "APPROVED": "อนุมัติ",
    "PARTIAL": "รับบางส่วน",   # เผื่ออนาคต
    "FULL": "รับครบ",          # เผื่ออนาคต
    "CANCELLED": "ยกเลิก",
}

PO_STATUS_TH_TO_LEGACY = {}
for k, v in PO_STATUS_LEGACY_TO_TH.items():
    PO_STATUS_TH_TO_LEGACY.setdefault(v, set()).add(k)


def _asset_status_display(s: str | None) -> str:
    """แสดงสถานะเป็นไทยสำหรับ UI (รองรับค่าอังกฤษเดิม)."""
    ss = (s or "").strip()
    if not ss:
        return ""
    up = ss.upper()
    if up in ASSET_STATUS_LEGACY_TO_TH:
        return ASSET_STATUS_LEGACY_TO_TH[up]
    return ss  # ถ้าเป็นไทยอยู่แล้วก็คืนค่าเดิม


def _po_status_display(s: str | None) -> str:
    """แสดงสถานะ PO เป็นไทยสำหรับ UI (รองรับค่าอังกฤษเดิม)."""
    ss = (s or "").strip()
    if not ss:
        return ""
    up = ss.upper()
    if up in PO_STATUS_LEGACY_TO_TH:
        return PO_STATUS_LEGACY_TO_TH[up]
    return ss


def _asset_status_filter_values(selected: str) -> list[str]:
    """
    คืนรายการค่าที่ต้องใช้ filter ใน DB:
    - ถ้าเลือกเป็นไทย => match ทั้งไทย และอังกฤษเดิมที่เกี่ยวข้อง
    - ถ้าเลือกเป็นอังกฤษเดิม => match เฉพาะอังกฤษเดิม + ไทยที่ map ได้
    """
    s = (selected or "").strip()
    if not s:
        return []
    up = s.upper()

    # เลือกด้วย legacy code
    if up in ASSET_STATUS_LEGACY_TO_TH:
        th = ASSET_STATUS_LEGACY_TO_TH[up]
        vals = {up, th}
        return list(vals)

    # เลือกด้วยไทย
    vals = {s}
    for legacy in ASSET_STATUS_TH_TO_LEGACY.get(s, set()):
        vals.add(legacy)
    return list(vals)


def _po_status_filter_values(selected: str) -> list[str]:
    s = (selected or "").strip()
    if not s:
        return []
    up = s.upper()

    # legacy
    if up in PO_STATUS_LEGACY_TO_TH:
        th = PO_STATUS_LEGACY_TO_TH[up]
        vals = {up, th}
        return list(vals)

    # thai
    vals = {s}
    for legacy in PO_STATUS_TH_TO_LEGACY.get(s, set()):
        vals.add(legacy)
    return list(vals)


# =========================================================
# Helpers
# =========================================================
def _to_date(s: str | None) -> date | None:
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except Exception:
        return None


def _gen_doc_no(prefix2: str, d: date, width: int = 5) -> str:
    """Generic doc no: PREFIX + YYMM + running.
    Example QT2603xxxxx, CT2603xxxxx, BL2603xxxxx, TX2603xxxxx, RC2603xxxxx
    """
    yymm = f"{d.year % 100:02d}{d.month:02d}"
    prefix = f"{prefix2}{yymm}"

    model_map = {
        "QT": Quote,
        "CT": RentalContract,
        "BL": Billing,
        "TX": TaxInvoice,
        "RC": Receipt,
    }
    m = model_map[prefix2]

    last = m.query.filter(m.doc_no.like(prefix + "%")).order_by(m.id.desc()).first()
    if not last:
        return prefix + ("0" * (width - 1)) + "1" if width == 1 else prefix + f"{1:0{width}d}"

    try:
        n = int(str(last.doc_no)[-width:]) + 1
    except Exception:
        n = last.id + 1

    return prefix + f"{n:0{width}d}"


def recalc_quote(q: Quote) -> None:
    lines_sum = Decimal("0")

    for ln in q.lines:
        qty = Decimal(str(ln.qty or 0))
        days = Decimal(str(ln.days or 0))
        price = Decimal(str(ln.price or 0))
        line_total = qty * days * price
        ln.line_total = line_total
        lines_sum += line_total

    q.deposit_extra = Decimal(str(q.deposit_extra or 0))

    vat_mode = (q.vat_mode or "EXCLUDED").upper()
    vat_rate = Decimal(str(q.vat_rate or 0)) / Decimal("100")

    if vat_mode == "NONE":
        subtotal = lines_sum
        vat = Decimal("0")
        incl = lines_sum
    elif vat_mode == "INCLUDED":
        incl = lines_sum
        subtotal = (incl / (Decimal("1") + vat_rate)) if vat_rate else incl
        vat = incl - subtotal
    else:  # EXCLUDED
        subtotal = lines_sum
        vat = subtotal * vat_rate
        incl = subtotal + vat

    wht_pct = Decimal(str(q.wht_percent or 0)) / Decimal("100")
    wht = subtotal * wht_pct
    net = incl - wht

    q.subtotal = subtotal
    q.vat_amount = vat
    q.total_incl_vat = incl
    q.wht_amount = wht
    q.net_payable = net


def _add_months(d: date, months: int) -> date:
    """Add months to date without external deps (clamps day to month end)."""
    y = d.year + (d.month - 1 + months) // 12
    m = (d.month - 1 + months) % 12 + 1
    if m in (1, 3, 5, 7, 8, 10, 12):
        last_day = 31
    elif m in (4, 6, 9, 11):
        last_day = 30
    else:
        leap = (y % 4 == 0 and (y % 100 != 0 or y % 400 == 0))
        last_day = 29 if leap else 28
    day = min(d.day, last_day)
    return date(y, m, day)


def _d(x) -> Decimal:
    try:
        return Decimal(str(x or 0))
    except Exception:
        return Decimal("0")


def _fmt_doc_type(th: str) -> str:
    mp = {
        "QT": "ใบเสนอราคา",
        "CT": "สัญญาเช่า",
        "BL": "ใบวางบิล/ใบแจ้งหนี้",
        "TX": "ใบกำกับภาษี",
        "RC": "ใบเสร็จรับเงิน",
    }
    return mp.get(th, th)


def _get_customer_installments(customer_id: int):
    """
    Return installments belonging to customer by join:
    Installment -> Contract -> Quote -> Customer
    """
    return (
        Installment.query.join(RentalContract, Installment.contract_id == RentalContract.id)
        .join(Quote, RentalContract.quote_id == Quote.id)
        .filter(Quote.customer_id == customer_id)
        .order_by(Installment.due_date.asc(), Installment.seq.asc(), Installment.id.asc())
        .all()
    )


def _inst_debit_source(ins: Installment):
    """
    Rule:
      - If TX exists and status == 'อนุมัติแล้ว' => debit = TX.amount (source=TX)
      - else if BL exists and status == 'วางบิลแล้ว' => debit = BL.amount (source=BL)
      - else debit = 0

    return (source_code, doc_obj_or_none, debit_decimal)
    """
    tx = getattr(ins, "tax_invoice", None)
    bl = getattr(ins, "billing", None)

    if tx and (tx.status or "") == "อนุมัติแล้ว":
        return ("TX", tx, _d(tx.amount))
    if bl and (bl.status or "") == "วางบิลแล้ว":
        return ("BL", bl, _d(bl.amount))
    return ("", None, Decimal("0"))


def _inst_paid(ins: Installment) -> Decimal:
    rc = getattr(ins, "receipt", None)
    if rc and (rc.status or "") == "รับเงินแล้ว":
        return _d(rc.amount)
    return Decimal("0")


def _build_customer_ar(customer_id: int):
    """
    Build:
      - ar_total (เกิดหนี้แล้ว)
      - paid_total (รับเงินแล้ว)
      - outstanding_total = ar_total - paid_total
      - outstanding_rows per installment
      - statement_rows timeline with running balance
    """
    installments = _get_customer_installments(customer_id)

    ar_total = Decimal("0")
    paid_total = Decimal("0")

    outstanding_rows = []
    events = []

    for ins in installments:
        src, doc_obj, debit = _inst_debit_source(ins)
        paid = _inst_paid(ins)

        if debit > 0:
            ar_total += debit
        if paid > 0:
            paid_total += paid

        # outstanding by installment
        out = debit - paid
        if out < 0:
            out = Decimal("0")

        outstanding_rows.append(
            {
                "installment": ins,
                "seq": getattr(ins, "seq", None),
                "due_date": getattr(ins, "due_date", None),
                "amount": _d(getattr(ins, "amount", None)),
                "debit_source": src,
                "debit_doc": doc_obj,
                "debit": debit,
                "receipt": getattr(ins, "receipt", None),
                "paid": paid,
                "outstanding": out,
            }
        )

        # statement events
        if debit > 0 and doc_obj:
            events.append(
                {
                    "doc_date": getattr(doc_obj, "doc_date", None) or getattr(ins, "due_date", None),
                    "type": src,
                    "type_name": _fmt_doc_type(src),
                    "doc_no": getattr(doc_obj, "doc_no", None),
                    "project_site": (
                        getattr(ins.contract, "project_site", None)
                        or getattr(ins.contract.quote, "project_site", None)
                    ),
                    "debit": debit,
                    "credit": Decimal("0"),
                    "ref": doc_obj,
                }
            )

        rc = getattr(ins, "receipt", None)
        if rc and (rc.status or "") == "รับเงินแล้ว":
            events.append(
                {
                    "doc_date": getattr(rc, "doc_date", None) or getattr(ins, "due_date", None),
                    "type": "RC",
                    "type_name": _fmt_doc_type("RC"),
                    "doc_no": getattr(rc, "doc_no", None),
                    "project_site": (
                        getattr(ins.contract, "project_site", None)
                        or getattr(ins.contract.quote, "project_site", None)
                    ),
                    "debit": Decimal("0"),
                    "credit": _d(getattr(rc, "amount", None)),
                    "ref": rc,
                }
            )

    # sort events by date; debit before credit on same day
    type_priority = {"BL": 10, "TX": 10, "RC": 20}
    events.sort(
        key=lambda e: (
            e["doc_date"] or date.min,
            type_priority.get(e["type"], 99),
            e.get("doc_no") or "",
        )
    )

    # running balance
    bal = Decimal("0")
    statement_rows = []
    for e in events:
        bal = bal + _d(e["debit"]) - _d(e["credit"])
        statement_rows.append({**e, "balance": bal})

    outstanding_total = ar_total - paid_total
    if outstanding_total < 0:
        outstanding_total = Decimal("0")

    return {
        "ar_total": ar_total,
        "paid_total": paid_total,
        "outstanding_total": outstanding_total,
        "outstanding_rows": outstanding_rows,
        "statement_rows": statement_rows,
    }


def _parse_filters():
    """
    Common report filters.
    Returns dict with:
      date_from, date_to, customer, project, doc, type, status
    """
    date_from = _to_date(request.args.get("date_from"))
    date_to = _to_date(request.args.get("date_to"))
    q_customer = (request.args.get("customer") or "").strip()
    q_project = (request.args.get("project") or "").strip()
    q_doc = (request.args.get("doc") or "").strip()
    dtype = (request.args.get("type") or "").strip().upper()
    status = (request.args.get("status") or "").strip()
    return {
        "date_from": date_from,
        "date_to": date_to,
        "customer": q_customer,
        "project": q_project,
        "doc": q_doc,
        "type": dtype,
        "status": status,
    }


def _in_date_range(d: date | None, date_from: date | None, date_to: date | None) -> bool:
    if not d:
        return False if (date_from or date_to) else True
    if date_from and d < date_from:
        return False
    if date_to and d > date_to:
        return False
    return True


# =========================================================
# Dashboard
# =========================================================
@bp.get("/")
def dashboard():
    return render_template("dashboard.html")


# =========================================================
# QUOTES
# =========================================================
@bp.get("/quotes")
def quote_list():
    rows = Quote.query.order_by(Quote.doc_date.desc(), Quote.id.desc()).all()
    return render_template("quotes/list.html", rows=rows)


@bp.get("/quotes/new")
def quote_new():
    customers = Customer.query.filter_by(is_active=True).order_by(Customer.name.asc()).all()
    products = Product.query.filter_by(is_active=True).order_by(Product.name.asc()).all()
    return render_template("quotes/form.html", mode="new", customers=customers, products=products, q=None)


@bp.post("/quotes/new")
def quote_create():
    customer_id = request.form.get("customer_id", type=int)
    if not customer_id:
        flash("กรุณาเลือกลูกค้า", "danger")
        return redirect(url_for("pages.quote_new"))

    d = _to_date(request.form.get("doc_date")) or date.today()

    q = Quote(
        doc_no=_gen_doc_no("QT", d),
        doc_date=d,
        customer_id=customer_id,
        rent_start=_to_date(request.form.get("rent_start")),
        rent_end=_to_date(request.form.get("rent_end")),
        project_site=request.form.get("project_site") or None,
        vat_mode=request.form.get("vat_mode") or "EXCLUDED",
        vat_rate=Decimal(request.form.get("vat_rate") or "7.00"),
        wht_percent=Decimal(request.form.get("wht_percent") or "0"),
        remark=request.form.get("remark") or None,
        status="ร่าง",
    )

    q.deposit_extra = Decimal(request.form.get("deposit_extra") or "0")

    db.session.add(q)
    db.session.flush()

    product_ids = request.form.getlist("product_id")
    qtys = request.form.getlist("qty")
    days_list = request.form.getlist("days")
    prices = request.form.getlist("price")
    deposits = request.form.getlist("deposit")

    for pid, qty, days, price, dep in zip(product_ids, qtys, days_list, prices, deposits):
        if not pid:
            continue
        db.session.add(
            QuoteLine(
                quote_id=q.id,
                product_id=int(pid),
                qty=Decimal(qty or "0"),
                days=int(days or "1"),
                price=Decimal(price or "0"),
                deposit=Decimal(dep or "0"),
            )
        )

    db.session.flush()
    recalc_quote(q)
    db.session.commit()

    flash("สร้างใบเสนอราคาแล้ว (ร่าง)", "success")
    return redirect(url_for("pages.quote_view", quote_id=q.id))


@bp.get("/quotes/<int:quote_id>")
def quote_view(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    return render_template("quotes/view.html", q=q)


@bp.get("/quotes/<int:quote_id>/installments")
def quote_installments(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    c = q.contract
    if not c:
        flash("ใบเสนอราคานี้ยังไม่มีสัญญาเช่า (CT) จึงยังไม่มีงวด", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))
    return render_template("quotes/installments.html", q=q, c=c)


@bp.get("/quotes/<int:quote_id>/docs")
def quote_docs(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    c = q.contract
    if not c:
        flash("ใบเสนอราคานี้ยังไม่มีสัญญาเช่า (CT) จึงยังไม่มีเอกสาร BL/TAX/RC", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))
    return render_template("quotes/docs.html", q=q, c=c)


@bp.get("/quotes/<int:quote_id>/edit")
def quote_edit(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    if q.status in ("อนุมัติ", "อนุมัติแล้ว"):
        flash("เอกสารถูกอนุมัติแล้ว ไม่สามารถแก้ไขได้", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))
    if q.status == "ยกเลิก":
        flash("เอกสารถูกยกเลิกแล้ว ไม่สามารถแก้ไขได้", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))

    customers = Customer.query.filter_by(is_active=True).order_by(Customer.name.asc()).all()
    products = Product.query.filter_by(is_active=True).order_by(Product.name.asc()).all()
    return render_template("quotes/form.html", mode="edit", customers=customers, products=products, q=q)


@bp.post("/quotes/<int:quote_id>/edit")
def quote_update(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    if q.status in ("อนุมัติ", "อนุมัติแล้ว"):
        flash("เอกสารถูกอนุมัติแล้ว ไม่สามารถแก้ไขได้", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))
    if q.status == "ยกเลิก":
        flash("เอกสารถูกยกเลิกแล้ว ไม่สามารถแก้ไขได้", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))

    q.deposit_extra = Decimal(request.form.get("deposit_extra") or "0")
    q.customer_id = request.form.get("customer_id", type=int) or q.customer_id
    q.doc_date = _to_date(request.form.get("doc_date")) or q.doc_date
    q.rent_start = _to_date(request.form.get("rent_start"))
    q.rent_end = _to_date(request.form.get("rent_end"))
    q.project_site = request.form.get("project_site") or None

    q.vat_mode = request.form.get("vat_mode") or "EXCLUDED"
    q.vat_rate = Decimal(request.form.get("vat_rate") or "7.00")
    q.wht_percent = Decimal(request.form.get("wht_percent") or "0")
    q.remark = request.form.get("remark") or None

    QuoteLine.query.filter_by(quote_id=q.id).delete()

    product_ids = request.form.getlist("product_id")
    qtys = request.form.getlist("qty")
    days_list = request.form.getlist("days")
    prices = request.form.getlist("price")
    deposits = request.form.getlist("deposit")

    for pid, qty, days, price, dep in zip(product_ids, qtys, days_list, prices, deposits):
        if not pid:
            continue
        db.session.add(
            QuoteLine(
                quote_id=q.id,
                product_id=int(pid),
                qty=Decimal(qty or "0"),
                days=int(days or "1"),
                price=Decimal(price or "0"),
                deposit=Decimal(dep or "0"),
            )
        )

    db.session.flush()
    recalc_quote(q)
    db.session.commit()

    flash("บันทึกการแก้ไขแล้ว", "success")
    return redirect(url_for("pages.quote_view", quote_id=q.id))


@bp.post("/quotes/<int:quote_id>/approve")
def quote_approve(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    if q.status in ("อนุมัติ", "อนุมัติแล้ว"):
        flash("เอกสารนี้อนุมัติแล้ว", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))
    if q.status == "ยกเลิก":
        flash("เอกสารถูกยกเลิกแล้ว", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))

    q.status = "อนุมัติแล้ว"
    q.approved_at = datetime.utcnow()
    db.session.commit()

    flash("อนุมัติเอกสารเรียบร้อยแล้ว", "success")
    return redirect(url_for("pages.quote_view", quote_id=q.id))


@bp.post("/quotes/<int:quote_id>/cancel")
def quote_cancel(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    if q.status == "ยกเลิก":
        flash("เอกสารนี้ถูกยกเลิกแล้ว", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))

    q.status = "ยกเลิก"
    q.canceled_at = datetime.utcnow()
    db.session.commit()

    flash("ยกเลิกเอกสารเรียบร้อยแล้ว", "success")
    return redirect(url_for("pages.quote_view", quote_id=q.id))


# =========================================================
# CONTRACT
# =========================================================
@bp.get("/quotes/<int:quote_id>/contract")
def contract_from_quote(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    if not q.is_approved:
        flash("ต้องอนุมัติใบเสนอราคาก่อน จึงจะสร้างสัญญาเช่าได้", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))

    if q.contract:
        return redirect(url_for("pages.contract_view", contract_id=q.contract.id))

    return render_template("contracts/create_from_quote.html", q=q)


@bp.post("/quotes/<int:quote_id>/contract")
def contract_create_from_quote(quote_id: int):
    q = Quote.query.get_or_404(quote_id)
    if not q.is_approved:
        flash("ต้องอนุมัติใบเสนอราคาก่อน จึงจะสร้างสัญญาเช่าได้", "warning")
        return redirect(url_for("pages.quote_view", quote_id=q.id))

    if q.contract:
        flash("ใบเสนอราคานี้มีสัญญาแล้ว", "warning")
        return redirect(url_for("pages.contract_view", contract_id=q.contract.id))

    d = date.today()
    c = RentalContract(
        quote_id=q.id,
        doc_no=_gen_doc_no("CT", d),
        doc_date=d,
        start_date=q.rent_start,
        end_date=q.rent_end,
        project_site=q.project_site,
        status="สัญญากำลังดำเนินการ",
        remark=q.remark,
    )

    db.session.add(c)
    db.session.commit()

    flash("สร้างสัญญาเช่าแล้ว (สัญญากำลังดำเนินการ)", "success")
    return redirect(url_for("pages.contract_view", contract_id=c.id))


@bp.get("/contracts/<int:contract_id>")
def contract_view(contract_id: int):
    c = RentalContract.query.get_or_404(contract_id)

    # ✅ normalize: ไม่ให้มีสถานะ "ร่าง"
    ensure_contract_running_status(c)

    # ✅ อัปเดตสถานะสัญญาอัตโนมัติทุกครั้งที่เปิดดู (กรณีข้อมูลเก่าชำระครบแล้ว)
    auto_update_contract_status(c)

    q = c.quote
    return render_template("contracts/view.html", c=c, q=q)


@bp.post("/contracts/<int:contract_id>/cancel")
def contract_cancel(contract_id: int):
    c = RentalContract.query.get_or_404(contract_id)

    if (c.status or "") == "ยกเลิกสัญญา":
        flash("สัญญานี้ถูกยกเลิกแล้ว", "warning")
        return redirect(url_for("pages.contract_view", contract_id=c.id))

    if (c.status or "") == "ปิดสัญญา":
        flash("สัญญานี้ปิดสัญญาแล้ว ไม่สามารถยกเลิกได้", "warning")
        return redirect(url_for("pages.contract_view", contract_id=c.id))

    c.status = "ยกเลิกสัญญา"
    db.session.commit()

    flash("ยกเลิกสัญญาเรียบร้อยแล้ว (ล็อกการทำงานทั้งหมด)", "success")
    return redirect(url_for("pages.contract_view", contract_id=c.id))


@bp.post("/contracts/<int:contract_id>/generate_installments")
def contract_generate_installments(contract_id: int):
    c = RentalContract.query.get_or_404(contract_id)
    q = c.quote

    resp = guard_contract_not_locked(c)
    if resp:
        return resp

    months = request.form.get("months", type=int) or 0
    if months <= 0:
        flash("กรุณาระบุจำนวนเดือน/จำนวนงวดให้ถูกต้อง", "danger")
        return redirect(url_for("pages.contract_view", contract_id=c.id))

    if c.installments:
        has_paid = any((ins.status or "") == "ชำระแล้ว" for ins in c.installments)
        has_docs = any((ins.billing or ins.tax_invoice or ins.receipt) for ins in c.installments)
        if has_paid or has_docs:
            flash("มีงวดที่ชำระแล้ว หรือมีเอกสารแล้ว ไม่อนุญาตให้สร้างงวดใหม่เพื่อป้องกันข้อมูลเพี้ยน", "warning")
            return redirect(url_for("pages.contract_view", contract_id=c.id))

    Installment.query.filter_by(contract_id=c.id).delete()
    db.session.flush()

    total = Decimal(str(q.net_payable or 0))
    if months == 1:
        amounts = [total]
    else:
        base = (total / Decimal(months)).quantize(Decimal("0.01"))
        amounts = [base for _ in range(months)]
        s = sum(amounts)
        diff = (total - s).quantize(Decimal("0.01"))
        amounts[-1] = (amounts[-1] + diff).quantize(Decimal("0.01"))

    start = c.start_date or q.rent_start or date.today()

    for i in range(months):
        due = _add_months(start, i)
        ins = Installment(
            contract_id=c.id,
            seq=i + 1,
            due_date=due,
            amount=amounts[i],
            status="ยังไม่ชำระ",
        )
        db.session.add(ins)

    db.session.commit()

    flash("สร้างงวดอัตโนมัติเรียบร้อยแล้ว", "success")
    return redirect(url_for("pages.contract_view", contract_id=c.id))


# =========================================================
# INSTALLMENT + DOCS
# =========================================================
@bp.get("/installments/<int:inst_id>")
def installment_view(inst_id: int):
    ins = Installment.query.get_or_404(inst_id)
    return render_template("installments/view.html", ins=ins)


@bp.post("/installments/<int:inst_id>/mark_paid")
def installment_mark_paid(inst_id: int):
    ins = Installment.query.get_or_404(inst_id)

    resp = guard_contract_not_locked(ins.contract)
    if resp:
        return resp

    if ins.status == "ชำระแล้ว":
        flash("งวดนี้ถูกชำระแล้ว", "warning")
        return redirect(url_for("pages.installment_view", inst_id=ins.id))

    ins.status = "ชำระแล้ว"
    ins.paid_at = datetime.utcnow()
    ins.paid_note = request.form.get("paid_note") or None
    ins.evidence_path = request.form.get("evidence_path") or None

    db.session.commit()

    # ✅ อัปเดตสถานะสัญญาอัตโนมัติ (idempotent)
    auto_update_contract_status(ins.contract)

    flash("บันทึกการชำระเงินแล้ว", "success")

    return redirect(url_for("pages.installment_view", inst_id=ins.id))


@bp.post("/installments/<int:inst_id>/mark_unpaid")
def installment_mark_unpaid(inst_id: int):
    ins = Installment.query.get_or_404(inst_id)

    resp = guard_contract_not_locked(ins.contract)
    if resp:
        return resp

    if ins.receipt and (ins.receipt.status == "รับเงินแล้ว"):
        flash("งวดนี้มีใบเสร็จรับเงินที่รับเงินแล้ว ไม่สามารถย้อนกลับเป็นยังไม่ชำระได้", "warning")
        return redirect(url_for("pages.installment_view", inst_id=ins.id))

    ins.status = "ยังไม่ชำระ"
    ins.paid_at = None
    ins.paid_note = None
    ins.evidence_path = None

    db.session.commit()
    flash("ปรับเป็นยังไม่ชำระแล้ว", "success")
    return redirect(url_for("pages.installment_view", inst_id=ins.id))


def _ensure_billing(ins: Installment) -> Billing:
    if ins.billing:
        return ins.billing
    d = date.today()
    b = Billing(
        installment_id=ins.id,
        doc_no=_gen_doc_no("BL", d),
        doc_date=d,
        amount=ins.amount,
        status="ยังไม่วางบิล",
    )
    db.session.add(b)
    db.session.commit()
    return b


def _ensure_tax(ins: Installment) -> TaxInvoice:
    if ins.tax_invoice:
        return ins.tax_invoice
    d = date.today()
    t = TaxInvoice(
        installment_id=ins.id,
        doc_no=_gen_doc_no("TX", d),
        doc_date=d,
        amount=ins.amount,
        status="ยังไม่อนุมัติ",
    )
    db.session.add(t)
    db.session.commit()
    return t


def _ensure_rc(ins: Installment) -> Receipt:
    if ins.receipt:
        return ins.receipt
    d = date.today()
    r = Receipt(
        installment_id=ins.id,
        doc_no=_gen_doc_no("RC", d),
        doc_date=d,
        amount=ins.amount,
        status="ยังไม่รับเงิน",
    )
    db.session.add(r)
    db.session.commit()
    return r


@bp.post("/installments/<int:inst_id>/create_billing")
def installment_create_billing(inst_id: int):
    ins = Installment.query.get_or_404(inst_id)

    resp = guard_contract_not_locked(ins.contract)
    if resp:
        return resp

    b = _ensure_billing(ins)
    flash("สร้างใบวางบิล/ใบแจ้งหนี้แล้ว", "success")
    return redirect(url_for("pages.billing_view", billing_id=b.id))


@bp.post("/installments/<int:inst_id>/create_tax")
def installment_create_tax(inst_id: int):
    ins = Installment.query.get_or_404(inst_id)

    resp = guard_contract_not_locked(ins.contract)
    if resp:
        return resp

    t = _ensure_tax(ins)
    flash("สร้างใบกำกับภาษีแล้ว", "success")
    return redirect(url_for("pages.tax_view", tax_id=t.id))


@bp.post("/installments/<int:inst_id>/create_receipt")
def installment_create_receipt(inst_id: int):
    ins = Installment.query.get_or_404(inst_id)

    resp = guard_contract_not_locked(ins.contract)
    if resp:
        return resp

    r = _ensure_rc(ins)
    flash("สร้างใบเสร็จรับเงินแล้ว", "success")
    return redirect(url_for("pages.receipt_view", receipt_id=r.id))


@bp.get("/billings/<int:billing_id>")
def billing_view(billing_id: int):
    b = Billing.query.get_or_404(billing_id)
    return render_template("docs/billing_view.html", b=b)


@bp.get("/tax/<int:tax_id>")
def tax_view(tax_id: int):
    t = TaxInvoice.query.get_or_404(tax_id)
    return render_template("docs/tax_view.html", t=t)


@bp.get("/receipts/<int:receipt_id>")
def receipt_view(receipt_id: int):
    r = Receipt.query.get_or_404(receipt_id)
    return render_template("docs/receipt_view.html", r=r)


# -------------------------
# ✅ PRINT ROUTES (A1)
# -------------------------
@bp.get("/billings/<int:billing_id>/print")
def billing_print(billing_id: int):
    b = Billing.query.get_or_404(billing_id)
    return render_template("print/billing_print.html", b=b)


@bp.get("/tax/<int:tax_id>/print")
def tax_print(tax_id: int):
    t = TaxInvoice.query.get_or_404(tax_id)
    return render_template("print/tax_print.html", t=t)


@bp.get("/receipts/<int:receipt_id>/print")
def receipt_print(receipt_id: int):
    r = Receipt.query.get_or_404(receipt_id)
    return render_template("print/receipt_print.html", r=r)


# -------------------------
# ✅ PDF ROUTES (A2: ReportLab)
# -------------------------
@bp.get("/billings/<int:billing_id>/pdf")
def billing_pdf(billing_id: int):
    b = Billing.query.get_or_404(billing_id)
    from app.utils.pdf_docs import build_billing_pdf

    pdf = build_billing_pdf(b)
    return send_file(
        BytesIO(pdf.content),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=pdf.filename,
    )


@bp.get("/tax/<int:tax_id>/pdf")
def tax_pdf(tax_id: int):
    t = TaxInvoice.query.get_or_404(tax_id)
    from app.utils.pdf_docs import build_tax_pdf

    pdf = build_tax_pdf(t)
    return send_file(
        BytesIO(pdf.content),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=pdf.filename,
    )


@bp.get("/receipts/<int:receipt_id>/pdf")
def receipt_pdf(receipt_id: int):
    r = Receipt.query.get_or_404(receipt_id)
    from app.utils.pdf_docs import build_receipt_pdf

    pdf = build_receipt_pdf(r)
    return send_file(
        BytesIO(pdf.content),
        mimetype="application/pdf",
        as_attachment=True,
        download_name=pdf.filename,
    )


@bp.post("/billings/<int:billing_id>/update_date")
def billing_update_date(billing_id: int):
    b = Billing.query.get_or_404(billing_id)

    resp = guard_contract_not_locked(b.installment.contract)
    if resp:
        return resp

    new_date = _to_date(request.form.get("doc_date"))
    if not new_date:
        flash("กรุณาเลือกวันที่เอกสารให้ถูกต้อง", "danger")
        return redirect(url_for("pages.billing_view", billing_id=b.id))

    b.doc_date = new_date
    b.amount = b.installment.amount
    db.session.commit()

    flash("บันทึกวันที่ใบวางบิล/ใบแจ้งหนี้แล้ว", "success")
    return redirect(url_for("pages.billing_view", billing_id=b.id))


@bp.post("/tax/<int:tax_id>/update_date")
def tax_update_date(tax_id: int):
    t = TaxInvoice.query.get_or_404(tax_id)

    resp = guard_contract_not_locked(t.installment.contract)
    if resp:
        return resp

    new_date = _to_date(request.form.get("doc_date"))
    if not new_date:
        flash("กรุณาเลือกวันที่เอกสารให้ถูกต้อง", "danger")
        return redirect(url_for("pages.tax_view", tax_id=t.id))

    t.doc_date = new_date
    t.amount = t.installment.amount
    db.session.commit()

    flash("บันทึกวันที่ใบกำกับภาษีแล้ว", "success")
    return redirect(url_for("pages.tax_view", tax_id=t.id))


@bp.post("/receipts/<int:receipt_id>/update_date")
def receipt_update_date(receipt_id: int):
    r = Receipt.query.get_or_404(receipt_id)

    resp = guard_contract_not_locked(r.installment.contract)
    if resp:
        return resp

    new_date = _to_date(request.form.get("doc_date"))
    if not new_date:
        flash("กรุณาเลือกวันที่เอกสารให้ถูกต้อง", "danger")
        return redirect(url_for("pages.receipt_view", receipt_id=r.id))

    r.doc_date = new_date
    r.amount = r.installment.amount
    db.session.commit()

    flash("บันทึกวันที่ใบเสร็จรับเงินแล้ว", "success")
    return redirect(url_for("pages.receipt_view", receipt_id=r.id))


# =========================================================
# CUSTOMERS (แฟ้มลูกค้า)
# =========================================================
@bp.get("/customers")
def customer_list():
    q = (request.args.get("q") or "").strip()

    qry = Customer.query
    if q:
        like = f"%{q}%"
        qry = qry.filter(
            or_(
                Customer.name.ilike(like),
                Customer.phone.ilike(like),
                Customer.tax_id.ilike(like),
                Customer.email.ilike(like),
                Customer.address.ilike(like),
            )
        )

    rows = qry.order_by(Customer.name.asc()).all()
    return render_template("customers/list.html", rows=rows, q=q)


@bp.get("/customers/<int:customer_id>")
def customer_view(customer_id: int):
    c = Customer.query.get_or_404(customer_id)

    project = (request.args.get("project") or "").strip()
    doc = (request.args.get("doc") or "").strip()
    dtype = (request.args.get("type") or "").strip().upper()

    like_project = f"%{project}%" if project else None
    like_doc = f"%{doc}%" if doc else None
    tab = (request.args.get("tab") or "t1").strip()

    # -------------------------
    # QT
    # -------------------------
    quotes = []
    if dtype in ("", "QT"):
        q_qry = Quote.query.filter(Quote.customer_id == c.id)
        if like_project:
            q_qry = q_qry.filter(Quote.project_site.ilike(like_project))
        if like_doc:
            q_qry = q_qry.filter(Quote.doc_no.ilike(like_doc))
        quotes = q_qry.order_by(Quote.doc_date.desc(), Quote.id.desc()).all()

    # -------------------------
    # CT
    # -------------------------
    contracts = []
    if dtype in ("", "CT"):
        ct_qry = RentalContract.query.join(Quote, RentalContract.quote_id == Quote.id).filter(
            Quote.customer_id == c.id
        )
        if like_project:
            ct_qry = ct_qry.filter(
                or_(
                    RentalContract.project_site.ilike(like_project),
                    Quote.project_site.ilike(like_project),
                )
            )
        if like_doc:
            ct_qry = ct_qry.filter(RentalContract.doc_no.ilike(like_doc))
        contracts = ct_qry.order_by(RentalContract.doc_date.desc(), RentalContract.id.desc()).all()

    # -------------------------
    # BL
    # -------------------------
    billings = []
    if dtype in ("", "BL"):
        b_qry = (
            Billing.query.join(Installment, Billing.installment_id == Installment.id)
            .join(RentalContract, Installment.contract_id == RentalContract.id)
            .join(Quote, RentalContract.quote_id == Quote.id)
            .filter(Quote.customer_id == c.id)
        )
        if like_project:
            b_qry = b_qry.filter(
                or_(
                    RentalContract.project_site.ilike(like_project),
                    Quote.project_site.ilike(like_project),
                )
            )
        if like_doc:
            b_qry = b_qry.filter(Billing.doc_no.ilike(like_doc))
        billings = b_qry.order_by(Billing.doc_date.desc(), Billing.id.desc()).all()

    # -------------------------
    # TX
    # -------------------------
    taxes = []
    if dtype in ("", "TX"):
        t_qry = (
            TaxInvoice.query.join(Installment, TaxInvoice.installment_id == Installment.id)
            .join(RentalContract, Installment.contract_id == RentalContract.id)
            .join(Quote, RentalContract.quote_id == Quote.id)
            .filter(Quote.customer_id == c.id)
        )
        if like_project:
            t_qry = t_qry.filter(
                or_(
                    RentalContract.project_site.ilike(like_project),
                    Quote.project_site.ilike(like_project),
                )
            )
        if like_doc:
            t_qry = t_qry.filter(TaxInvoice.doc_no.ilike(like_doc))
        taxes = t_qry.order_by(TaxInvoice.doc_date.desc(), TaxInvoice.id.desc()).all()

    # -------------------------
    # RC
    # -------------------------
    receipts = []
    if dtype in ("", "RC"):
        r_qry = (
            Receipt.query.join(Installment, Receipt.installment_id == Installment.id)
            .join(RentalContract, Installment.contract_id == RentalContract.id)
            .join(Quote, RentalContract.quote_id == Quote.id)
            .filter(Quote.customer_id == c.id)
        )
        if like_project:
            r_qry = r_qry.filter(
                or_(
                    RentalContract.project_site.ilike(like_project),
                    Quote.project_site.ilike(like_project),
                )
            )
        if like_doc:
            r_qry = r_qry.filter(Receipt.doc_no.ilike(like_doc))
        receipts = r_qry.order_by(Receipt.doc_date.desc(), Receipt.id.desc()).all()

    # -------------------------
    # ✅ AR / Statement / Outstanding (คำนวณจริง)
    # -------------------------
    ar_pack = _build_customer_ar(c.id)

    return render_template(
        "customers/view.html",
        c=c,
        quotes=quotes,
        contracts=contracts,
        billings=billings,
        taxes=taxes,
        receipts=receipts,
        tab=tab,
        project=project,
        doc=doc,
        dtype=dtype,
        ar_total=ar_pack["ar_total"],
        paid_total=ar_pack["paid_total"],
        outstanding_total=ar_pack["outstanding_total"],
        statement_rows=ar_pack["statement_rows"],
        outstanding_rows=ar_pack["outstanding_rows"],
    )


@bp.post("/billings/<int:billing_id>/toggle_status")
def billing_toggle_status(billing_id: int):
    b = Billing.query.get_or_404(billing_id)

    resp = guard_contract_not_locked(b.installment.contract)
    if resp:
        return resp

    if b.status == "วางบิลแล้ว":
        flash("ใบวางบิลนี้ถูกวางบิลแล้ว ไม่สามารถย้อนกลับได้", "warning")
        return redirect(url_for("pages.billing_view", billing_id=b.id))

    b.status = "วางบิลแล้ว"
    db.session.commit()

    flash("อัปเดตสถานะใบวางบิลเป็น 'วางบิลแล้ว' เรียบร้อย", "success")
    return redirect(url_for("pages.billing_view", billing_id=b.id))


@bp.post("/tax/<int:tax_id>/approve")
def tax_approve(tax_id: int):
    t = TaxInvoice.query.get_or_404(tax_id)

    resp = guard_contract_not_locked(t.installment.contract)
    if resp:
        return resp

    if t.status == "อนุมัติแล้ว":
        flash("ใบกำกับภาษีนี้อนุมัติแล้ว ไม่สามารถย้อนกลับได้", "warning")
        return redirect(url_for("pages.tax_view", tax_id=t.id))

    t.status = "อนุมัติแล้ว"
    db.session.commit()

    flash("อนุมัติใบกำกับภาษีแล้ว", "success")
    return redirect(url_for("pages.tax_view", tax_id=t.id))


@bp.post("/receipts/<int:receipt_id>/receive")
def receipt_receive(receipt_id: int):
    r = Receipt.query.get_or_404(receipt_id)

    resp = guard_contract_not_locked(r.installment.contract)
    if resp:
        return resp

    if r.status == "รับเงินแล้ว":
        flash("ใบเสร็จรับเงินนี้รับเงินแล้ว ไม่สามารถย้อนกลับได้", "warning")
        return redirect(url_for("pages.receipt_view", receipt_id=r.id))

    r.status = "รับเงินแล้ว"

    ins = r.installment
    if ins:
        ins.status = "ชำระแล้ว"
        if not ins.paid_at:
            ins.paid_at = datetime.utcnow()
        if not ins.paid_note:
            ins.paid_note = f"รับเงินผ่าน RC {r.doc_no}"

    db.session.commit()

    # ✅ อัปเดตสถานะสัญญาอัตโนมัติ (idempotent)
    auto_update_contract_status(r.installment.contract)

    flash("บันทึกรับเงินเรียบร้อยแล้ว และปรับงวดเป็น 'ชำระแล้ว' อัตโนมัติ", "success")
    return redirect(url_for("pages.receipt_view", receipt_id=r.id))


# =========================================================
# REPORT CENTER + EXPORT EXCEL
# =========================================================
def _collect_report_rows(filters: dict):
    """
    Collect rows across documents (QT/CT/BL/TX/RC) into unified dict rows,
    then filter in Python by date range / customer / project / doc / type / status.
    Returns list[dict].
    """
    date_from = filters["date_from"]
    date_to = filters["date_to"]
    f_customer = (filters["customer"] or "").strip()
    f_project = (filters["project"] or "").strip()
    f_doc = (filters["doc"] or "").strip()
    dtype = (filters["type"] or "").strip().upper()
    f_status = (filters["status"] or "").strip()

    like_customer = f_customer.lower()
    like_project = f_project.lower()
    like_doc = f_doc.lower()
    like_status = f_status.lower()

    rows = []

    def ok_text(hay: str | None, needle_lower: str) -> bool:
        if not needle_lower:
            return True
        return needle_lower in (hay or "").lower()

    # QT
    if dtype in ("", "QT"):
        for q in Quote.query.order_by(Quote.doc_date.desc(), Quote.id.desc()).all():
            if not _in_date_range(getattr(q, "doc_date", None), date_from, date_to):
                continue
            cust_name = q.customer.name if q.customer else ""
            proj = q.project_site or ""
            if not ok_text(cust_name, like_customer):
                continue
            if not ok_text(proj, like_project):
                continue
            if not ok_text(q.doc_no, like_doc):
                continue
            if like_status and like_status not in (q.status or "").lower():
                continue

            rows.append(
                {
                    "type": "QT",
                    "type_name": _fmt_doc_type("QT"),
                    "doc_date": q.doc_date,
                    "doc_no": q.doc_no,
                    "status": q.status,
                    "customer_name": cust_name,
                    "project_site": proj,
                    "amount": _d(getattr(q, "net_payable", None)),
                    "link": url_for("pages.quote_view", quote_id=q.id),
                }
            )

    # CT
    if dtype in ("", "CT"):
        ct_qry = (
            RentalContract.query.join(Quote, RentalContract.quote_id == Quote.id)
            .join(Customer, Quote.customer_id == Customer.id)
            .order_by(RentalContract.doc_date.desc(), RentalContract.id.desc())
        )
        for c in ct_qry.all():
            q = c.quote
            cust_name = q.customer.name if q and q.customer else ""
            proj = c.project_site or (q.project_site if q else "") or ""
            if not _in_date_range(getattr(c, "doc_date", None), date_from, date_to):
                continue
            if not ok_text(cust_name, like_customer):
                continue
            if not ok_text(proj, like_project):
                continue
            if not ok_text(c.doc_no, like_doc):
                continue
            if like_status and like_status not in (c.status or "").lower():
                continue

            rows.append(
                {
                    "type": "CT",
                    "type_name": _fmt_doc_type("CT"),
                    "doc_date": c.doc_date,
                    "doc_no": c.doc_no,
                    "status": c.status,
                    "customer_name": cust_name,
                    "project_site": proj,
                    "amount": _d(getattr(q, "net_payable", None)) if q else Decimal("0"),
                    "link": url_for("pages.contract_view", contract_id=c.id),
                }
            )

    # BL
    if dtype in ("", "BL"):
        b_qry = (
            Billing.query.join(Installment, Billing.installment_id == Installment.id)
            .join(RentalContract, Installment.contract_id == RentalContract.id)
            .join(Quote, RentalContract.quote_id == Quote.id)
            .join(Customer, Quote.customer_id == Customer.id)
            .order_by(Billing.doc_date.desc(), Billing.id.desc())
        )
        for b in b_qry.all():
            q = b.installment.contract.quote
            cust_name = q.customer.name if q.customer else ""
            proj = b.installment.contract.project_site or q.project_site or ""
            if not _in_date_range(getattr(b, "doc_date", None), date_from, date_to):
                continue
            if not ok_text(cust_name, like_customer):
                continue
            if not ok_text(proj, like_project):
                continue
            if not ok_text(b.doc_no, like_doc):
                continue
            if like_status and like_status not in (b.status or "").lower():
                continue

            rows.append(
                {
                    "type": "BL",
                    "type_name": _fmt_doc_type("BL"),
                    "doc_date": b.doc_date,
                    "doc_no": b.doc_no,
                    "status": b.status,
                    "customer_name": cust_name,
                    "project_site": proj,
                    "amount": _d(b.amount),
                    "link": url_for("pages.billing_view", billing_id=b.id),
                }
            )

    # TX
    if dtype in ("", "TX"):
        t_qry = (
            TaxInvoice.query.join(Installment, TaxInvoice.installment_id == Installment.id)
            .join(RentalContract, Installment.contract_id == RentalContract.id)
            .join(Quote, RentalContract.quote_id == Quote.id)
            .join(Customer, Quote.customer_id == Customer.id)
            .order_by(TaxInvoice.doc_date.desc(), TaxInvoice.id.desc())
        )
        for t in t_qry.all():
            q = t.installment.contract.quote
            cust_name = q.customer.name if q.customer else ""
            proj = t.installment.contract.project_site or q.project_site or ""
            if not _in_date_range(getattr(t, "doc_date", None), date_from, date_to):
                continue
            if not ok_text(cust_name, like_customer):
                continue
            if not ok_text(proj, like_project):
                continue
            if not ok_text(t.doc_no, like_doc):
                continue
            if like_status and like_status not in (t.status or "").lower():
                continue

            rows.append(
                {
                    "type": "TX",
                    "type_name": _fmt_doc_type("TX"),
                    "doc_date": t.doc_date,
                    "doc_no": t.doc_no,
                    "status": t.status,
                    "customer_name": cust_name,
                    "project_site": proj,
                    "amount": _d(t.amount),
                    "link": url_for("pages.tax_view", tax_id=t.id),
                }
            )

    # RC
    if dtype in ("", "RC"):
        r_qry = (
            Receipt.query.join(Installment, Receipt.installment_id == Installment.id)
            .join(RentalContract, Installment.contract_id == RentalContract.id)
            .join(Quote, RentalContract.quote_id == Quote.id)
            .join(Customer, Quote.customer_id == Customer.id)
            .order_by(Receipt.doc_date.desc(), Receipt.id.desc())
        )
        for r in r_qry.all():
            q = r.installment.contract.quote
            cust_name = q.customer.name if q.customer else ""
            proj = r.installment.contract.project_site or q.project_site or ""
            if not _in_date_range(getattr(r, "doc_date", None), date_from, date_to):
                continue
            if not ok_text(cust_name, like_customer):
                continue
            if not ok_text(proj, like_project):
                continue
            if not ok_text(r.doc_no, like_doc):
                continue
            if like_status and like_status not in (r.status or "").lower():
                continue

            rows.append(
                {
                    "type": "RC",
                    "type_name": _fmt_doc_type("RC"),
                    "doc_date": r.doc_date,
                    "doc_no": r.doc_no,
                    "status": r.status,
                    "customer_name": cust_name,
                    "project_site": proj,
                    "amount": _d(r.amount),
                    "link": url_for("pages.receipt_view", receipt_id=r.id),
                }
            )

    # Sort newest first for UI
    rows.sort(key=lambda x: (x["doc_date"] or date.min, x["doc_no"] or ""), reverse=True)
    return rows


@bp.get("/reports")
def reports_index():
    filters = _parse_filters()
    rows = _collect_report_rows(filters)
    return render_template("reports/index.html", rows=rows, filters=filters)


@bp.get("/reports/export.xlsx")
def reports_export_xlsx():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    filters = _parse_filters()
    rows = _collect_report_rows(filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    headers = [
        "ประเภท",
        "วันที่",
        "เลขที่เอกสาร",
        "สถานะ",
        "ลูกค้า",
        "โครงการ/หน้างาน",
        "ยอดเงิน",
        "ลิงก์",
    ]
    ws.append(headers)

    for r in rows:
        ws.append(
            [
                r.get("type_name"),
                (r.get("doc_date").isoformat() if r.get("doc_date") else ""),
                r.get("doc_no") or "",
                r.get("status") or "",
                r.get("customer_name") or "",
                r.get("project_site") or "",
                float(_d(r.get("amount"))),
                r.get("link") or "",
            ]
        )

    # basic column widths
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 32

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = "report_export.xlsx"
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# =========================================================
# PURCHASE / INVENTORY PAGES (BUY & STOCK)
# =========================================================
def _gen_po_no(d: date) -> str:
    yymm = f"{d.year % 100:02d}{d.month:02d}"
    prefix = f"PO{yymm}"
    last = PurchaseOrder.query.filter(PurchaseOrder.po_no.like(prefix + "%")).order_by(PurchaseOrder.id.desc()).first()
    if not last:
        return prefix + f"{1:05d}"
    try:
        n = int(str(last.po_no)[-5:]) + 1
    except Exception:
        n = last.id + 1
    return prefix + f"{n:05d}"


def _gen_grn_no(d: date) -> str:
    yymm = f"{d.year % 100:02d}{d.month:02d}"
    prefix = f"GR{yymm}"
    last = GRN.query.filter(GRN.grn_no.like(prefix + "%")).order_by(GRN.id.desc()).first()
    if not last:
        return prefix + f"{1:05d}"
    try:
        n = int(str(last.grn_no)[-5:]) + 1
    except Exception:
        n = last.id + 1
    return prefix + f"{n:05d}"


@bp.get("/suppliers")
def supplier_list():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()  # active/inactive/blank

    qry = Supplier.query
    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(Supplier.name.ilike(like), Supplier.phone.ilike(like), Supplier.tax_id.ilike(like)))

    if status == "active":
        qry = qry.filter(Supplier.is_active.is_(True))
    elif status == "inactive":
        qry = qry.filter(Supplier.is_active.is_(False))

    rows = qry.order_by(Supplier.id.desc()).limit(500).all()
    return render_template(
        "suppliers/list.html",
        q=q,
        status=status,
        rows=rows,
    )


@bp.get("/purchase/po")
def po_page():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    supplier_id = request.args.get("supplier_id", type=int)

    qry = PurchaseOrder.query.join(Supplier, PurchaseOrder.supplier_id == Supplier.id)

    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(PurchaseOrder.po_no.ilike(like), Supplier.name.ilike(like)))

    # ✅ filter รองรับทั้งสถานะไทย/อังกฤษเดิม
    if status:
        vals = _po_status_filter_values(status)
        if vals:
            qry = qry.filter(PurchaseOrder.status.in_(vals))
        else:
            qry = qry.filter(PurchaseOrder.status == status)

    if supplier_id:
        qry = qry.filter(PurchaseOrder.supplier_id == supplier_id)

    rows = qry.order_by(PurchaseOrder.id.desc()).limit(300).all()

    # ✅ ให้ template เอาไปแสดงสถานะไทยได้เลย ถ้าจะใช้
    # (ไม่บังคับ template ต้องแก้ทันที แต่พร้อมรองรับ)
    for po in rows:
        setattr(po, "_status_th", _po_status_display(getattr(po, "status", None)))

    suppliers = Supplier.query.filter(Supplier.is_active.is_(True)).order_by(Supplier.name.asc()).all()
    cats = Category.query.filter(Category.is_active.is_(True)).order_by(Category.name.asc()).all()
    units = Unit.query.filter(Unit.is_active.is_(True)).order_by(Unit.name.asc()).all()

    return render_template(
        "purchase/po.html",
        q=q,
        status=status,
        supplier_id=supplier_id,
        rows=rows,
        suppliers=suppliers,
        categories=cats,
        units=units,
        today=date.today(),
        po_statuses_th=PO_STATUSES_TH,
    )


@bp.get("/purchase/grn")
def grn_page():
    q = (request.args.get("q") or "").strip()
    supplier_id = request.args.get("supplier_id", type=int)
    mode = (request.args.get("mode") or "po").strip()  # po/cash

    qry = GRN.query.join(Supplier, GRN.supplier_id == Supplier.id)
    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(GRN.grn_no.ilike(like), Supplier.name.ilike(like)))
    if supplier_id:
        qry = qry.filter(GRN.supplier_id == supplier_id)

    rows = qry.order_by(GRN.id.desc()).limit(300).all()

    suppliers = Supplier.query.filter(Supplier.is_active.is_(True)).order_by(Supplier.name.asc()).all()

    # ✅ เลือก PO สำหรับ GRN: รองรับทั้ง "APPROVED" เดิม และ "อนุมัติ" ใหม่
    approved_vals = list({*PO_STATUS_TH_TO_LEGACY.get("อนุมัติ", set()), "อนุมัติ", "APPROVED"})
    pos = (
        PurchaseOrder.query.filter(PurchaseOrder.status.in_(approved_vals))
        .order_by(PurchaseOrder.id.desc())
        .limit(300)
        .all()
    )

    cats = Category.query.filter(Category.is_active.is_(True)).order_by(Category.name.asc()).all()
    units = Unit.query.filter(Unit.is_active.is_(True)).order_by(Unit.name.asc()).all()
    brands = Brand.query.filter(Brand.is_active.is_(True)).order_by(Brand.name.asc()).all()

    # หมายเหตุ: print นี้เอาออกได้ แต่คงไว้เพื่อ debug เดิม
    print("APPROVED PO count =", len(pos), [p.po_no for p in pos[:5]])

    # ✅ ใส่ field ช่วย template
    for po in pos:
        setattr(po, "_status_th", _po_status_display(getattr(po, "status", None)))

    return render_template(
        "purchase/grn.html",
        q=q,
        supplier_id=supplier_id,
        mode=mode,
        rows=rows,
        suppliers=suppliers,
        pos=pos,
        categories=cats,
        units=units,
        brands=brands,
        today=date.today(),
    )


@bp.get("/stock/dashboard")
def stock_dashboard():
    # ✅ counts by status (รองรับไทย + อังกฤษเดิม)
    statuses = ASSET_STATUSES_TH[:]  # ไทยล้วนสำหรับ UI
    counts = {s: 0 for s in statuses}

    for s in statuses:
        vals = _asset_status_filter_values(s)
        if vals:
            counts[s] = Asset.query.filter(Asset.status.in_(vals)).count()
        else:
            counts[s] = Asset.query.filter(Asset.status == s).count()

    # summary by category
    cats = Category.query.order_by(Category.name.asc()).all()
    cat_rows = []
    for c in cats:
        total = Asset.query.filter(Asset.category_id == c.id).count()

        def _cnt(status_th: str) -> int:
            vals = _asset_status_filter_values(status_th)
            if vals:
                return Asset.query.filter(Asset.category_id == c.id, Asset.status.in_(vals)).count()
            return Asset.query.filter(Asset.category_id == c.id, Asset.status == status_th).count()

        avail = _cnt("พร้อมใช้งาน")
        reserved = _cnt("จองแล้ว")
        rented = _cnt("กำลังเช่า")
        inspect = _cnt("รอตรวจสภาพ")
        repair = _cnt("ซ่อม")
        lost = _cnt("สูญหาย")
        retired = _cnt("จำหน่ายออก")

        cat_rows.append(
            {
                "cat": c,
                "total": total,
                "available": avail,
                "reserved": reserved,
                "rented": rented,
                "inspect": inspect,
                "repair": repair,
                "lost": lost,
                "retired": retired,
            }
        )

    return render_template("stock/dashboard.html", counts=counts, cat_rows=cat_rows, asset_statuses_th=ASSET_STATUSES_TH)


@bp.get("/stock/categories")
def category_page():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()

    qry = Category.query
    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(Category.code.ilike(like), Category.name.ilike(like)))

    if status == "active":
        qry = qry.filter(Category.is_active.is_(True))
    elif status == "inactive":
        qry = qry.filter(Category.is_active.is_(False))

    rows = qry.order_by(Category.id.desc()).limit(500).all()
    return render_template("stock/categories.html", q=q, status=status, rows=rows)


@bp.get("/stock/assets")
def asset_page():
    q = (request.args.get("q") or "").strip()
    category_id = request.args.get("category_id", type=int)
    status = (request.args.get("status") or "").strip()

    qry = Asset.query.join(Category, Asset.category_id == Category.id)

    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(Asset.asset_code.ilike(like), Asset.name.ilike(like), Category.name.ilike(like)))

    if category_id:
        qry = qry.filter(Asset.category_id == category_id)

    # ✅ filter รองรับไทย/อังกฤษเดิม
    if status:
        vals = _asset_status_filter_values(status)
        if vals:
            qry = qry.filter(Asset.status.in_(vals))
        else:
            qry = qry.filter(Asset.status == status)

    rows = qry.order_by(Asset.id.desc()).limit(800).all()

    # ✅ เติม field ช่วย template ให้แสดงไทยได้ทันที
    for a in rows:
        setattr(a, "_status_th", _asset_status_display(getattr(a, "status", None)))

    cats = Category.query.filter(Category.is_active.is_(True)).order_by(Category.name.asc()).all()
    suppliers = Supplier.query.filter(Supplier.is_active.is_(True)).order_by(Supplier.name.asc()).all()
    brands = Brand.query.filter(Brand.is_active.is_(True)).order_by(Brand.name.asc()).all()
    units = Unit.query.filter(Unit.is_active.is_(True)).order_by(Unit.name.asc()).all()

    return render_template(
        "stock/assets.html",
        q=q,
        category_id=category_id,
        status=status,
        rows=rows,
        categories=cats,
        suppliers=suppliers,
        brands=brands,
        units=units,
        today=date.today(),
        asset_statuses_th=ASSET_STATUSES_TH,
    )


@bp.get("/stock/adjust")
def stock_adjust_page():
    q = (request.args.get("q") or "").strip()
    category_id = request.args.get("category_id", type=int)
    status = (request.args.get("status") or "").strip()

    qry = Asset.query.join(Category, Asset.category_id == Category.id)

    if q:
        like = f"%{q}%"
        qry = qry.filter(or_(Asset.asset_code.ilike(like), Asset.name.ilike(like), Category.name.ilike(like)))

    if category_id:
        qry = qry.filter(Asset.category_id == category_id)

    # ✅ filter รองรับไทย/อังกฤษเดิม
    if status:
        vals = _asset_status_filter_values(status)
        if vals:
            qry = qry.filter(Asset.status.in_(vals))
        else:
            qry = qry.filter(Asset.status == status)

    rows = qry.order_by(Asset.id.desc()).limit(800).all()

    # ✅ เติม field ช่วย template
    for a in rows:
        setattr(a, "_status_th", _asset_status_display(getattr(a, "status", None)))

    cats = Category.query.filter(Category.is_active.is_(True)).order_by(Category.name.asc()).all()

    return render_template(
        "stock/adjust.html",
        q=q,
        category_id=category_id,
        status=status,
        rows=rows,
        categories=cats,
        asset_statuses_th=ASSET_STATUSES_TH,
    )